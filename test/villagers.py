import requests
from bs4 import BeautifulSoup
import re
import openpyxl

url = "https://nookipedia.com/wiki/List_of_villager_names_in_other_languages"

wb = openpyxl.Workbook()
ws = wb.create_sheet('villagers')

ws['A1'] = 'Eng'
ws['B1'] = 'Kor'




request = requests.get(url)
soup = BeautifulSoup(request.text, "html.parser")
table = soup.find('table', {'class': 'styled color-villager'})

trs = table.find_all('tr')
for idx, tr in enumerate(trs):
    tds = tr.find_all('td')
    for index, td in enumerate(tds):
        if index == 0:
            eng = re.sub(r"[^a-zA-Z]", "", td.text.strip())
            ws['A'+str(idx+1)] = eng
            print(eng, end=" ")
        if index == 6:
            kor = re.sub(r"[^가-힣]", "", td.text.strip())
            ws['B' + str(idx+1)] = kor

wb.save('/Users/SSAFY/acnh_kor.xlsx')

