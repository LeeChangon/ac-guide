import requests
from bs4 import BeautifulSoup
import re
import openpyxl

url = "https://nookipedia.com/wiki/Bug/New_Horizons"

fpath = '/Users/SSAFY/acnh_kor.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb.create_sheet('bugs')

ws['A1'] = 'Eng'
ws['B1'] = 'Kor'

request = requests.get(url)
soup = BeautifulSoup(request.text, "html.parser")
table = soup.find('table', {'class': 'sortable'})
trs = table.find_all('tr')
for idx, tr in enumerate(trs):
    tds = tr.find_all('td')
    for index, td in enumerate(tds):
        if index == 1:
            origin = re.sub(r"[^a-zA-Z\s\'\-]", "", td.text.strip())
            eng = origin.replace(" ", "_")
            eng = origin.replace("\'", "%27")
            nurl = "https://nookipedia.com/wiki/"+eng
            nrequest = requests.get(nurl)
            nsoup = BeautifulSoup(nrequest.text, "html.parser")
            data = nsoup.find('div', {'class': 'mw-collapsible-content'})
            print(origin)
            contents = data.find_all('span')
            for content in contents:
                kor = re.sub(r"[^가-힣]", "", content.text)
                if len(kor) != 0:
                    ws['A' + str(idx + 1)] = origin
                    ws['B' + str(idx + 1)] = kor
                    break;

wb.save('/Users/SSAFY/acnh_kor.xlsx')

